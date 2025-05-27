import openpyxl

# ─────────────────────────────────────────────────────────────
#  ORIGINAL FUNCTION — Base Excel handler
# ─────────────────────────────────────────────────────────────

def extract_text_from_excel(file_path):
    """
    Extracts readable text from an Excel (.xlsx) file.

    - Iterates through all worksheets
    - Reads all rows and cells
    - Handles empty cells gracefully
    - Formats output with tabs for clarity

    Args:
        file_path (str): Path to the .xlsx file.

    Returns:
        str: Cleaned, structured text from all sheets.
    """
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        extracted = []

        for sheet in workbook.worksheets:
            extracted.append(f"### Sheet: {sheet.title} ###")

            for row in sheet.iter_rows():
                # Convert all cell values to strings (empty -> "")
                row_text = [str(cell.value).strip() if cell.value is not None else "" for cell in row]

                # Skip empty rows
                if any(cell for cell in row_text):
                    extracted.append("\t".join(row_text))

            extracted.append("")  # Add a line break between sheets

        return "\n".join(extracted).strip()

    except Exception as e:
        print(f"[ERROR] Failed to extract XLSX: {file_path} – {e}")
        return ""

# ─────────────────────────────────────────────────────────────
#  BONUS FUNCTION — Enhanced Excel extraction with metadata
# ─────────────────────────────────────────────────────────────

import logging

# Logger setup
logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format='[%(levelname)s] %(message)s')

def extract_text_from_excel_enhanced(file_path, include_empty=False):
    """
    Enhanced extraction from Excel (.xlsx) files with optional control over
    empty rows and structured logging of sheet/cell statistics.

    Args:
        file_path (str): Path to the .xlsx file.
        include_empty (bool): If True, include empty rows; otherwise, skip them.

    Returns:
        str: Tabular, cleanly formatted text from all sheets.
    """
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet_count = len(workbook.worksheets)
        extracted = []

        logger.info(f"Workbook opened: {sheet_count} sheet(s) found.")

        for sheet_index, sheet in enumerate(workbook.worksheets, start=1):
            extracted.append(f"### Sheet {sheet_index}: {sheet.title} ###")
            row_counter = 0

            for row in sheet.iter_rows():
                row_text = [str(cell.value).strip() if cell.value is not None else "" for cell in row]

                if not any(row_text) and not include_empty:
                    continue  # Skip empty rows if not requested

                extracted.append("\t".join(row_text))
                row_counter += 1

            logger.info(f"Processed Sheet {sheet_index} – Rows extracted: {row_counter}")
            extracted.append("")  # Line break between sheets

        return "\n".join(extracted).strip()

    except Exception as e:
        logger.error(f"Enhanced Excel extraction failed for {file_path}: {e}")
        return ""
